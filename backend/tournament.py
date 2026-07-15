import io
import json
import random
import secrets
from copy import copy
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote

import pandas as pd
import qrcode
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont

from .constants import (
    AGE_GROUPS,
    FINAL_ROUNDS,
    GAME_ALIASES,
    GAME_PANDA,
    GAME_RAINFOREST,
    GAME_SNOW,
    GAMES,
    IDENTITY_FREE_AGENT,
    IDENTITY_NORMAL,
    IDENTITY_ORIGINAL_SCHOOL,
    PRELIM_ROUNDS,
    SCORE_MAP,
    TRANSITION_CHOICES,
)
from .database import BASE_DIR, get_conn, rows_to_dicts


QR_DIR = BASE_DIR / "static" / "qrcodes"
PLAYER_QR_DIR = BASE_DIR / "static" / "player_qrcodes"
EXPORT_DIR = BASE_DIR / "exports"
TEMPLATE_DIR = BASE_DIR / "templates"
LIVE_SCORE_TEMPLATE = TEMPLATE_DIR / "成绩统计表模板.xlsx"
QR_FONT_PATHS = [
    "/System/Library/Fonts/PingFang.ttc",
    "/System/Library/Fonts/STHeiti Light.ttc",
    "/Library/Fonts/Arial Unicode.ttf",
]


def identity_label(is_transition: bool, transition_choice: str | None) -> str:
    if not is_transition:
        return IDENTITY_NORMAL
    if transition_choice == "自由人补位":
        return IDENTITY_FREE_AGENT
    return IDENTITY_ORIGINAL_SCHOOL


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if pd.isna(value):
        return False
    text = str(value).strip().lower()
    return text in {"是", "true", "1", "yes", "y", "升学", "自由人"}


def normalize_choice(value: Any, is_transition: bool) -> str | None:
    if not is_transition:
        return None
    if pd.isna(value) or str(value).strip() == "":
        raise ValueError("升学过渡选手必须填写升学出路选择")
    text = str(value).strip()
    aliases = {
        "自由人": "自由人补位",
        "自由人补位": "自由人补位",
        "跟随原校": "跟随原学校",
        "跟随原学校": "跟随原学校",
        "原校特批": "跟随原学校",
    }
    choice = aliases.get(text)
    if choice not in TRANSITION_CHOICES:
        raise ValueError(f"升学出路选择只能是：{'、'.join(TRANSITION_CHOICES)}")
    return choice


def clean_cell(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def load_excel_table(content: bytes) -> pd.DataFrame:
    raw = pd.read_excel(io.BytesIO(content), header=None)
    header_row = None
    for index, row in raw.iterrows():
        values = {clean_cell(value) for value in row.tolist()}
        if "序号" in values and ("队名" in values or "队伍名称" in values):
            header_row = index
            break
        if "姓名" in values and "队伍名称" in values:
            header_row = index
            break
    if header_row is None:
        return pd.read_excel(io.BytesIO(content))
    columns = [clean_cell(value) or f"未命名列{col_index + 1}" for col_index, value in enumerate(raw.iloc[header_row].tolist())]
    frame = raw.iloc[header_row + 1 :].copy()
    frame.columns = columns
    frame = frame.dropna(how="all")
    return frame


def first_value(row: pd.Series, names: list[str], default: str = "") -> str:
    for name in names:
        if name in row.index:
            value = clean_cell(row[name])
            if value:
                return value
    return default


def has_team_roster_columns(frame: pd.DataFrame) -> bool:
    columns = set(frame.columns)
    if {"区", "学校", "队名"}.issubset(columns):
        return True
    team_name_cols = {"队名", "队伍名称"}
    member_cols = {"队员1姓名", "姓名1", "队员1", "队员1名"}
    return bool(team_name_cols & columns) and bool(member_cols & columns)


def normalize_age_group(value: str) -> str:
    text = clean_cell(value)
    if text in AGE_GROUPS:
        return text
    if "低" in text or "2-3" in text or "1-3" in text:
        return "小学低年级组"
    if "高" in text or "4-5" in text:
        return "小学高年级组"
    if "初中" in text:
        return "初中组"
    return text


def normalize_game_type(value: str) -> str:
    text = clean_cell(value)
    if text in GAME_ALIASES:
        return GAME_ALIASES[text]
    if text in GAMES:
        return text
    if "熊猫" in text:
        return GAME_PANDA
    if "雪山" in text or "三江源" in text:
        return GAME_SNOW
    if "雨林" in text:
        return GAME_RAINFOREST
    return text


def player_code_prefix(age_group: str) -> str:
    return {"小学低年级组": "A", "小学高年级组": "B", "初中组": "C"}.get(age_group, "P")


def roster_member_value(row: pd.Series, index: int, kind: str) -> str:
    if kind == "name":
        aliases = [
            f"队员{index}姓名",
            f"队员{index}名",
            f"姓名{index}",
            f"队员{index}",
            f"队员{index}姓名/编号",
        ]
    else:
        aliases = [
            f"队员{index}编号",
            f"编号{index}",
            f"队员{index}号码",
            f"队员{index}参赛编号",
            f"选手{index}编号",
        ]
    return first_value(row, aliases)


def records_from_team_roster(frame: pd.DataFrame) -> tuple[list[dict], list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    records: list[dict] = []
    seed_rank = 1
    generated_codes = 0
    code_counters: dict[str, int] = defaultdict(int)

    if not any(column in frame.columns for column in ["年级组别", "年龄组别", "组别"]):
        warnings.append("队伍名单表未包含“年级组别”，已默认按“小学低年级组”导入。")
    if not any(column in frame.columns for column in ["所属游戏", "游戏类型", "项目", "挑战内容"]):
        warnings.append(f"队伍名单表未包含“所属游戏”，已默认按“{GAMES[0]}”导入。")

    for row_index, raw in frame.iterrows():
        row_no = row_index + 2
        try:
            team_name = first_value(raw, ["队名", "队伍名称"])
            member_names = [roster_member_value(raw, member_index, "name") for member_index in range(1, 5)]
            if not team_name and not any(member_names):
                continue
            if not team_name:
                raise ValueError("队名不能为空")
            district = first_value(raw, ["区", "区县", "区域"])
            school = first_value(raw, ["学校", "学校名称"])
            team_code = first_value(raw, ["队伍编号", "队编号", "团队编号", "编号", "序号"])
            age_group = normalize_age_group(first_value(raw, ["年级组别", "年龄组别", "组别"], AGE_GROUPS[0]))
            game_type = normalize_game_type(first_value(raw, ["所属游戏", "游戏类型", "项目", "挑战内容"], GAMES[0]))
            is_transition = parse_bool(first_value(raw, ["是否升学过渡"], "否"))
            choice = normalize_choice(first_value(raw, ["升学出路选择"]), is_transition)
            if age_group not in AGE_GROUPS:
                raise ValueError(f"年级组别必须是：{'、'.join(AGE_GROUPS)}")
            if game_type not in GAMES:
                raise ValueError(f"所属游戏必须是：{'、'.join(GAMES)}")

            members: list[tuple[str, str]] = []
            for member_index in range(1, 5):
                member_name = member_names[member_index - 1]
                player_code = roster_member_value(raw, member_index, "code")
                if not member_name:
                    raise ValueError(f"队员{member_index}姓名不能为空")
                if not player_code:
                    prefix = player_code_prefix(age_group)
                    code_counters[prefix] += 1
                    player_code = f"{prefix}{code_counters[prefix]:03d}"
                    generated_codes += 1
                members.append((member_name, player_code))

            for member_name, player_code in members:
                records.append(
                    {
                        "name": member_name,
                        "team_name": team_name,
                        "district": district,
                        "school": school,
                        "team_code": team_code,
                        "player_code": player_code,
                        "age_group": age_group,
                        "game_type": game_type,
                        "seed_rank": seed_rank,
                        "is_transition": int(is_transition),
                        "transition_choice": choice,
                        "identity_label": identity_label(is_transition, choice),
                    }
                )
                seed_rank += 1
        except Exception as exc:
            errors.append(f"第{row_no}行：{exc}")

    if generated_codes:
        warnings.append(f"名单未提供队员编号，系统已按组别自动生成 {generated_codes} 个编号（低年级A、高年级B、初中C）。")
    return records, errors, warnings


def team_match_key(player: dict) -> str:
    return clean_cell(player.get("team_code")) or f"{clean_cell(player.get('school'))}|{clean_cell(player.get('team_name'))}|{clean_cell(player.get('game_type'))}"


def records_from_player_rows(frame: pd.DataFrame) -> tuple[list[dict], list[str], list[str]]:
    required = ["姓名", "队伍名称", "年级组别", "所属游戏", "赛前种子排名", "是否升学过渡"]
    missing = [name for name in required if name not in frame.columns]
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少必填列：{', '.join(missing)}")

    errors: list[str] = []
    warnings: list[str] = []
    records: list[dict] = []

    for index, raw in frame.iterrows():
        row_no = index + 2
        try:
            name = clean_cell(raw["姓名"])
            team = clean_cell(raw["队伍名称"])
            age_group = normalize_age_group(clean_cell(raw["年级组别"]))
            game_type = normalize_game_type(clean_cell(raw["所属游戏"]))
            seed_rank = int(raw["赛前种子排名"])
            is_transition = parse_bool(raw["是否升学过渡"])
            choice_value = raw["升学出路选择"] if "升学出路选择" in frame.columns else None
            choice = normalize_choice(choice_value, is_transition)
            if not name:
                raise ValueError("姓名不能为空")
            if not team:
                raise ValueError("队伍名称不能为空")
            if age_group not in AGE_GROUPS:
                raise ValueError(f"年级组别必须是：{'、'.join(AGE_GROUPS)}")
            if game_type not in GAMES:
                raise ValueError(f"所属游戏必须是：{'、'.join(GAMES)}")
            records.append(
                {
                    "name": name,
                    "team_name": team,
                    "district": first_value(raw, ["区", "区县", "区域"]),
                    "school": first_value(raw, ["学校", "学校名称"]),
                    "team_code": first_value(raw, ["队伍编号", "队编号", "团队编号"]),
                    "player_code": first_value(raw, ["队员编号", "选手编号", "个人编号", "编号"]),
                    "age_group": age_group,
                    "game_type": game_type,
                    "seed_rank": seed_rank,
                    "is_transition": int(is_transition),
                    "transition_choice": choice,
                    "identity_label": identity_label(is_transition, choice),
                }
            )
        except Exception as exc:
            errors.append(f"第{row_no}行：{exc}")
    return records, errors, warnings


def import_players_from_excel(content: bytes, force: bool, base_url: str = "") -> dict:
    try:
        frame = load_excel_table(content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel读取失败：{exc}") from exc

    if has_team_roster_columns(frame):
        records, errors, warnings = records_from_team_roster(frame)
        import_format = "队伍名单表"
    else:
        records, errors, warnings = records_from_player_rows(frame)
        import_format = "选手明细表"

    if not records:
        errors.append("Excel中没有可导入的选手")

    seed_ranks = [record["seed_rank"] for record in records]
    if len(seed_ranks) != len(set(seed_ranks)):
        errors.append("赛前种子排名存在重复")
    if seed_ranks and sorted(seed_ranks) != list(range(1, len(seed_ranks) + 1)):
        errors.append("赛前种子排名必须是从1开始连续且不缺失的整数")
    player_codes = [record["player_code"] for record in records if record.get("player_code")]
    if len(player_codes) != len(set(player_codes)):
        errors.append("队员编号存在重复，请检查名单")

    pool_counter: dict[tuple[str, str], int] = defaultdict(int)
    team_counter: dict[tuple[str, str], int] = defaultdict(int)
    for record in records:
        if record["identity_label"] != IDENTITY_FREE_AGENT:
            pool_counter[(record["age_group"], record["game_type"])] += 1
            team_counter[(team_match_key(record), record["game_type"])] += 1

    for (age_group, game_type), count in sorted(pool_counter.items()):
        if count % 4 != 0:
            warnings.append(f"{age_group} / {game_type} 普通匹配池人数为 {count}，不是4的倍数，后续可能产生3人桌、2人轮空或人工介入提醒。")
    for (team, game_type), count in sorted(team_counter.items()):
        if count != 4:
            warnings.append(f"队伍「{team}」在 {game_type} 下有 {count} 名非自由人成员，不等于固定4人。")

    if errors:
        raise HTTPException(status_code=400, detail={"errors": errors, "warnings": warnings})
    if warnings and not force:
        return {"imported": False, "warnings": warnings, "message": "发现警告，确认后可强制导入。"}

    with get_conn() as conn:
        conn.execute("BEGIN IMMEDIATE")
        conn.execute("DELETE FROM scores")
        conn.execute("DELETE FROM tables")
        conn.execute("DELETE FROM players")
        conn.execute("UPDATE rounds SET status = '未开始'")
        for record in records:
            cursor = conn.execute(
                """
                INSERT INTO players(
                    name, team_name, district, school, team_code, player_code,
                    age_group, game_type, seed_rank, is_active,
                    is_transition, transition_choice, identity_label
                ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?)
                """,
                (
                    record["name"],
                    record["team_name"],
                    record.get("district"),
                    record.get("school"),
                    record.get("team_code"),
                    record.get("player_code"),
                    record["age_group"],
                    record["game_type"],
                    record["seed_rank"],
                    record["is_transition"],
                    record["transition_choice"],
                    record["identity_label"],
                ),
            )
            player_id = int(cursor.lastrowid)
            token = create_player_qrcode(player_id, base_url, record.get("player_code"), record["name"])
            conn.execute("UPDATE players SET player_qr_token = ? WHERE id = ?", (token, player_id))
        conn.execute("COMMIT")

    return {"imported": True, "count": len(records), "warnings": warnings, "format": import_format}


def load_qr_font(size: int) -> ImageFont.ImageFont:
    for path in QR_FONT_PATHS:
        if Path(path).exists():
            return ImageFont.truetype(path, size=size)
    return ImageFont.load_default()


def text_width(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> int:
    bbox = draw.textbbox((0, 0), text, font=font)
    return bbox[2] - bbox[0]


def fit_qr_font(draw: ImageDraw.ImageDraw, text: str, max_width: int, start_size: int) -> ImageFont.ImageFont:
    for size in range(start_size, 11, -1):
        font = load_qr_font(size)
        if text_width(draw, text, font) <= max_width:
            return font
    return load_qr_font(11)


def draw_centered_text(
    draw: ImageDraw.ImageDraw,
    text: str,
    y: int,
    max_width: int,
    start_size: int,
    fill: str = "#111827",
) -> None:
    text = clean_cell(text) or "未填写"
    font = fit_qr_font(draw, text, max_width, start_size)
    x = (360 - text_width(draw, text, font)) // 2
    draw.text((x, y), text, font=font, fill=fill)


def save_labeled_qrcode(url: str, path: Path, player_name: str, player_code: str | None) -> None:
    qr = qrcode.QRCode(border=2, box_size=10)
    qr.add_data(url)
    qr.make(fit=True)
    qr_image = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    qr_image = qr_image.resize((300, 300), Image.Resampling.NEAREST)

    canvas = Image.new("RGB", (360, 410), "white")
    canvas.paste(qr_image, (30, 12))

    draw = ImageDraw.Draw(canvas)
    draw_centered_text(draw, player_name, 318, 320, 28)
    draw_centered_text(draw, clean_cell(player_code) or "无编号", 358, 320, 24, "#374151")
    canvas.save(path)


def safe_filename_part(value: Any, fallback: str) -> str:
    text = clean_cell(value) or fallback
    for char in r'\/:*?"<>|':
        text = text.replace(char, "_")
    return "_".join(text.split()) or fallback


def player_qr_filename(player_id: int, player_name: str = "", player_code: str | None = None) -> str:
    code = safe_filename_part(player_code, f"ID{player_id}")
    name = safe_filename_part(player_name, "未命名")
    return f"{code}_{name}.png"


def create_player_qrcode(
    player_id: int,
    base_url: str,
    player_code: str | None = None,
    player_name: str = "",
) -> str:
    from .security import make_player_token

    PLAYER_QR_DIR.mkdir(parents=True, exist_ok=True)
    token = make_player_token(player_id, player_code)
    host = base_url.rstrip("/") if base_url else "http://127.0.0.1:8000"
    player_url = f"{host}/player?token={token}"
    save_labeled_qrcode(player_url, PLAYER_QR_DIR / player_qr_filename(player_id, player_name, player_code), player_name, player_code)
    return token


def player_qr_public_path(player_id: int, player_name: str = "", player_code: str | None = None) -> str:
    if not player_name and not player_code:
        with get_conn() as conn:
            row = conn.execute("SELECT name, player_code FROM players WHERE id = ?", (player_id,)).fetchone()
        if row:
            player_name = row["name"]
            player_code = row["player_code"]
    qr_file = PLAYER_QR_DIR / player_qr_filename(player_id, player_name, player_code)
    version = qr_file.stat().st_mtime_ns if qr_file.exists() else int(datetime.now().timestamp() * 1000)
    return f"/static/player_qrcodes/{quote(qr_file.name)}?v={version}"


def refresh_player_qrcodes(base_url: str) -> dict:
    from .security import make_player_token

    PLAYER_QR_DIR.mkdir(parents=True, exist_ok=True)
    refreshed = 0
    with get_conn() as conn:
        rows = conn.execute("SELECT * FROM players WHERE is_active = 1 ORDER BY seed_rank").fetchall()
        active_filenames = {player_qr_filename(row["id"], row["name"], row["player_code"]) for row in rows}
        conn.execute("BEGIN IMMEDIATE")
        for row in rows:
            token = row["player_qr_token"] or make_player_token(row["id"], row["player_code"])
            player_url = f"{base_url.rstrip('/')}/player?token={token}"
            save_labeled_qrcode(player_url, PLAYER_QR_DIR / player_qr_filename(row["id"], row["name"], row["player_code"]), row["name"], row["player_code"])
            if not row["player_qr_token"]:
                conn.execute("UPDATE players SET player_qr_token = ? WHERE id = ?", (token, row["id"]))
            refreshed += 1
        conn.execute("COMMIT")

    removed = 0
    for qr_file in PLAYER_QR_DIR.glob("*.png"):
        if qr_file.name not in active_filenames:
            qr_file.unlink()
            removed += 1
    return {"refreshed": refreshed, "removed_old_files": removed}


def list_players(transition_only: bool = False) -> list[dict]:
    where = "WHERE is_transition = 1" if transition_only else ""
    with get_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT id, name, team_name, district, school, team_code, player_code,
                   player_qr_token,
                   age_group, game_type, seed_rank, is_active,
                   is_transition, transition_choice, identity_label
            FROM players
            {where}
            ORDER BY seed_rank ASC
            """
        ).fetchall()
    players = rows_to_dicts(rows)
    for player in players:
        player["player_qr_path"] = player_qr_public_path(player["id"], player["name"], player["player_code"])
    return players


def update_transition(players: list[dict]) -> dict:
    with get_conn() as conn:
        conn.execute("BEGIN IMMEDIATE")
        for item in players:
            player_id = int(item["id"])
            choice = item["transition_choice"]
            if choice not in TRANSITION_CHOICES:
                raise HTTPException(status_code=400, detail="升学出路选择无效")
            label = identity_label(True, choice)
            conn.execute(
                """
                UPDATE players
                SET is_transition = 1, transition_choice = ?, identity_label = ?
                WHERE id = ?
                """,
                (choice, label, player_id),
            )
        conn.execute("COMMIT")
    return {"updated": len(players)}


def get_round_id(conn, round_no: int) -> int:
    row = conn.execute("SELECT id FROM rounds WHERE round_no = ?", (round_no,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="轮次不存在")
    return int(row["id"])


def eligible_players(conn, round_no: int) -> list[dict]:
    if round_no == 1:
        rows = conn.execute("SELECT * FROM players WHERE is_active = 1 ORDER BY seed_rank").fetchall()
        return rows_to_dicts(rows)

    if round_no in {2, 3, 4, 5, 6}:
        previous = round_no - 1
        rows = conn.execute(
            """
            SELECT p.*
            FROM players p
            JOIN scores s ON s.player_id = p.id
            JOIN rounds r ON r.id = s.round_id
            WHERE p.is_active = 1 AND r.round_no = ? AND s.is_advanced = 1
            ORDER BY p.seed_rank
            """,
            (previous,),
        ).fetchall()
        return rows_to_dicts(rows)

    raise HTTPException(status_code=400, detail="轮次编号必须为1-6")


def seed_group(player: dict) -> int:
    return int(player.get("_seed_group", (int(player["seed_rank"]) - 1) // 4))


def take_distinct_seed_table(players: list[dict], size: int = 4) -> list[dict]:
    chosen: list[dict] = []
    used_groups: set[int] = set()
    for player in sorted(players, key=lambda item: int(item["seed_rank"])):
        group = seed_group(player)
        if group in used_groups:
            continue
        chosen.append(player)
        used_groups.add(group)
        if len(chosen) == size:
            break
    if len(chosen) < size:
        return []
    chosen_ids = {player["id"] for player in chosen}
    players[:] = [player for player in players if player["id"] not in chosen_ids]
    return chosen


def take_plain_table(players: list[dict], size: int = 4) -> list[dict]:
    table = players[:size]
    del players[:size]
    return table


def pick_team_safe_table(
    players: list[dict],
    size: int,
    enforce_seed: bool = True,
    preserve_order: bool = False,
) -> tuple[list[dict], bool]:
    chosen: list[dict] = []
    used_teams: set[str] = set()
    used_seed_groups: set[int] = set()
    had_team_conflict = False

    def candidates() -> list[dict]:
        if preserve_order:
            return list(players)
        return sorted(players, key=lambda item: (seed_group(item), int(item["seed_rank"])))

    def choose(require_team: bool, require_seed: bool) -> bool:
        for player in candidates():
            team = team_match_key(player)
            if require_team and team in used_teams:
                continue
            if require_seed and seed_group(player) in used_seed_groups:
                continue
            chosen.append(player)
            used_teams.add(team)
            used_seed_groups.add(seed_group(player))
            players.remove(player)
            return True
        return False

    if enforce_seed:
        while len(chosen) < size and choose(require_team=True, require_seed=True):
            pass
        while len(chosen) < size and choose(require_team=True, require_seed=False):
            pass
        while len(chosen) < size and choose(require_team=False, require_seed=True):
            had_team_conflict = True
            pass
        while len(chosen) < size and choose(require_team=False, require_seed=False):
            had_team_conflict = True
            pass
    else:
        while len(chosen) < size and choose(require_team=True, require_seed=False):
            pass
        while len(chosen) < size and choose(require_team=False, require_seed=False):
            had_team_conflict = True
            pass
    return chosen, had_team_conflict


def pair_key(left: int, right: int) -> tuple[int, int]:
    return (left, right) if left < right else (right, left)


def previous_final_opponent_pairs(conn, round_no: int) -> set[tuple[int, int]]:
    if round_no not in FINAL_ROUNDS:
        return set()
    rows = conn.execute(
        """
        SELECT t.player_ids_json
        FROM tables t
        JOIN rounds r ON r.id = t.round_id
        WHERE r.round_no >= 4 AND r.round_no < ?
        """,
        (round_no,),
    ).fetchall()
    pairs: set[tuple[int, int]] = set()
    for row in rows:
        ids = [int(player_id) for player_id in json.loads(row["player_ids_json"])]
        for index, player_id in enumerate(ids):
            for other_id in ids[index + 1 :]:
                pairs.add(pair_key(player_id, other_id))
    return pairs


def pick_final_table(players: list[dict], size: int, avoid_pairs: set[tuple[int, int]]) -> tuple[list[dict], bool]:
    chosen: list[dict] = []
    used_teams: set[str] = set()
    had_team_conflict = False
    while len(chosen) < size and players:
        best_index = 0
        best_score: tuple[int, int, int] | None = None
        for index, player in enumerate(players):
            player_id = int(player["id"])
            team_conflict = 1 if team_match_key(player) in used_teams else 0
            repeat_count = sum(1 for chosen_player in chosen if pair_key(player_id, int(chosen_player["id"])) in avoid_pairs)
            score = (team_conflict, repeat_count, index)
            if best_score is None or score < best_score:
                best_score = score
                best_index = index
        player = players.pop(best_index)
        if team_match_key(player) in used_teams:
            had_team_conflict = True
        chosen.append(player)
        used_teams.add(team_match_key(player))
    return chosen, had_team_conflict


def table_repeat_pair_count(table_players: list[dict], avoid_pairs: set[tuple[int, int]]) -> int:
    count = 0
    for index, player in enumerate(table_players):
        player_id = int(player["id"])
        for other_player in table_players[index + 1 :]:
            if pair_key(player_id, int(other_player["id"])) in avoid_pairs:
                count += 1
    return count


def final_plan_score(tables: list[dict], avoid_pairs: set[tuple[int, int]]) -> tuple[int, int]:
    repeat_pairs = sum(table_repeat_pair_count(table["players"], avoid_pairs) for table in tables)
    team_conflicts = sum(1 for table in tables if has_team_conflict(table["players"]))
    return repeat_pairs, team_conflicts


def build_final_tables_once(players: list[dict], avoid_pairs: set[tuple[int, int]]) -> list[dict]:
    remaining = list(players)
    tables: list[dict] = []
    while len(remaining) >= 4:
        table_players, had_conflict = pick_final_table(remaining, 4, avoid_pairs)
        tables.append({"players": table_players, "is_bye": False})
    if len(remaining) == 3:
        table_players, had_conflict = pick_final_table(remaining, 3, avoid_pairs)
        tables.append({"players": table_players, "is_bye": False})
    elif len(remaining) == 2:
        table_players, had_conflict = pick_final_table(remaining, 2, avoid_pairs)
        tables.append({"players": table_players, "is_bye": True})
    return tables


def build_best_final_tables(ordered_pool: list[dict], round_no: int, age_group: str, game_type: str, avoid_pairs: set[tuple[int, int]]) -> list[dict]:
    best_tables: list[dict] = []
    best_score: tuple[int, int] | None = None
    attempt_count = 240 if len(ordered_pool) >= 8 else 40
    base_ids = ",".join(str(player["id"]) for player in ordered_pool)
    for attempt in range(attempt_count):
        remaining = list(ordered_pool)
        seed_text = f"{round_no}|{age_group}|{game_type}|{base_ids}|{attempt}"
        random.Random(seed_text).shuffle(remaining)
        tables = build_final_tables_once(remaining, avoid_pairs)
        score = final_plan_score(tables, avoid_pairs)
        if best_score is None or score < best_score:
            best_tables = tables
            best_score = score
            if score == (0, 0):
                break
    return best_tables


def has_team_conflict(players: list[dict]) -> bool:
    keys = [team_match_key(player) for player in players]
    return len(keys) != len(set(keys))


def repair_team_conflicts(tables: list[dict]) -> None:
    for _ in range(200):
        changed = False
        for table in tables:
            players = table["players"]
            if not has_team_conflict(players):
                continue
            counts = Counter(team_match_key(player) for player in players)
            duplicate_keys = {key for key, count in counts.items() if count > 1}
            for player_index, player in enumerate(players):
                player_key = team_match_key(player)
                if player_key not in duplicate_keys:
                    continue
                table_keys_without_player = {team_match_key(item) for idx, item in enumerate(players) if idx != player_index}
                for other_table in tables:
                    if other_table is table:
                        continue
                    other_players = other_table["players"]
                    other_keys = {team_match_key(item) for item in other_players}
                    if player_key in other_keys:
                        continue
                    for other_index, other_player in enumerate(other_players):
                        other_key = team_match_key(other_player)
                        if other_key in table_keys_without_player:
                            continue
                        other_keys_without_player = {team_match_key(item) for idx, item in enumerate(other_players) if idx != other_index}
                        if other_key in other_keys_without_player:
                            continue
                        players[player_index], other_players[other_index] = other_players[other_index], players[player_index]
                        changed = True
                        break
                    if changed:
                        break
                if changed:
                    break
            if changed:
                break
        if not changed:
            break


def make_regular_tables(
    pool: list[dict],
    round_no: int,
    warnings: list[str],
    age_group: str,
    game_type: str,
    avoid_pairs: set[tuple[int, int]] | None = None,
) -> list[dict]:
    avoid_pairs = avoid_pairs or set()
    ordered_pool = sorted(pool, key=lambda item: int(item["seed_rank"]))
    for index, player in enumerate(ordered_pool):
        player["_seed_group"] = index // 4

    tables: list[dict] = []
    remaining = list(ordered_pool)

    if round_no in PRELIM_ROUNDS:
        while len(remaining) >= 4:
            table_players, had_conflict = pick_team_safe_table(remaining, 4)
            if len(table_players) < 4:
                remaining = table_players + remaining
                break
            tables.append({"players": table_players, "is_bye": False})
    else:
        tables = build_best_final_tables(ordered_pool, round_no, age_group, game_type, avoid_pairs)
        remaining = []

    if len(remaining) == 3:
        table_players, had_conflict = pick_team_safe_table(remaining, 3)
        tables.append({"players": table_players, "is_bye": False})
    elif len(remaining) == 2:
        table_players, had_conflict = pick_team_safe_table(remaining, 2)
        tables.append({"players": table_players, "is_bye": True})
    elif len(remaining) == 1:
        warnings.append(f"{age_group} / {game_type} 剩余1人无法成赛，需要管理员人工介入：{remaining[0]['name']}。")

    repair_team_conflicts(tables)
    for table in tables:
        if has_team_conflict(table["players"]):
            names = "、".join(player["name"] for player in table["players"])
            table_type = "2人轮空桌" if table["is_bye"] else f"{len(table['players'])}人桌"
            warnings.append(f"{age_group} / {game_type} {table_type}因人数结构限制出现同队同桌：{names}。")
    return tables


def build_table_plan(conn, round_no: int) -> tuple[list[dict], list[str]]:
    players = eligible_players(conn, round_no)
    avoid_pairs = previous_final_opponent_pairs(conn, round_no)
    regular_pools: dict[tuple[str, str], list[dict]] = defaultdict(list)
    free_agents_by_game: dict[str, list[dict]] = defaultdict(list)
    warnings: list[str] = []

    for player in players:
        if player["identity_label"] == IDENTITY_FREE_AGENT:
            free_agents_by_game[player["game_type"]].append(player)
        else:
            regular_pools[(player["age_group"], player["game_type"])].append(player)
    for game_type in free_agents_by_game:
        free_agents_by_game[game_type].sort(key=lambda item: int(item["seed_rank"]))

    planned: list[dict] = []
    for age_group in AGE_GROUPS:
        for game_type in GAMES:
            pool = regular_pools.get((age_group, game_type), [])
            tables = make_regular_tables(pool, round_no, warnings, age_group, game_type, avoid_pairs)
            for table in tables:
                planned.append(
                    {
                        "age_group": age_group,
                        "game_type": game_type,
                        "players": table["players"],
                        "is_bye": table["is_bye"],
                        "has_free_agent": False,
                    }
                )

    for game_type in GAMES:
        free_pool = free_agents_by_game.get(game_type, [])
        if not free_pool:
            continue
        tables = make_regular_tables(free_pool, round_no, warnings, "自由人组", game_type, avoid_pairs)
        for table in tables:
            planned.append(
                {
                    "age_group": "自由人组",
                    "game_type": game_type,
                    "players": table["players"],
                    "is_bye": table["is_bye"],
                    "has_free_agent": True,
                }
            )

    return planned, warnings


def insert_score(
    conn,
    player_id: int,
    round_id: int,
    table_id: int,
    rank: int | None,
    score: int,
    absent: bool,
    advanced: bool,
) -> None:
    conn.execute(
        """
        INSERT OR REPLACE INTO scores(
            player_id, round_id, table_id, table_rank, round_score, is_absent, is_advanced
        ) VALUES(?, ?, ?, ?, ?, ?, ?)
        """,
        (player_id, round_id, table_id, rank, score, int(absent), int(advanced)),
    )


def is_player_advanced(round_no: int, participant_count: int, is_bye: bool, rank: int, absent: bool) -> bool:
    if absent:
        return False
    if is_bye:
        return True
    if round_no in PRELIM_ROUNDS:
        return rank <= 2
    if round_no == 4:
        return rank == 1
    return False


def recompute_round5_advancement(conn, round_id: int) -> None:
    rows = conn.execute(
        """
        SELECT s.id AS score_id, s.player_id, s.round_score, s.table_rank, s.is_absent,
               p.seed_rank, t.age_group, t.game_type, t.submitted_at
        FROM scores s
        JOIN players p ON p.id = s.player_id
        JOIN tables t ON t.id = s.table_id
        WHERE s.round_id = ?
        """,
        (round_id,),
    ).fetchall()
    pools: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in rows:
        item = dict(row)
        pools[(item["age_group"], item["game_type"])].append(item)

    conn.execute("UPDATE scores SET is_advanced = 0 WHERE round_id = ?", (round_id,))
    for pool_rows in pools.values():
        ranked = sorted(
            pool_rows,
            key=lambda item: (
                int(item["is_absent"]),
                -int(item["round_score"]),
                item["submitted_at"] or "",
                int(item["table_rank"] or 999),
                int(item["seed_rank"]),
            ),
        )
        for item in ranked[:4]:
            if not item["is_absent"]:
                conn.execute("UPDATE scores SET is_advanced = 1 WHERE id = ?", (item["score_id"],))


def generate_round(round_no: int, base_url: str) -> dict:
    from .security import make_table_token

    with get_conn() as conn:
        round_id = get_round_id(conn, round_no)
        submitted_count = conn.execute(
            """
            SELECT COUNT(*) AS count
            FROM tables
            WHERE round_id = ? AND submitted_at IS NOT NULL
            """,
            (round_id,),
        ).fetchone()["count"]
        if submitted_count:
            raise HTTPException(status_code=400, detail="本轮已有成绩提交，不能重新生成桌位")

        conn.execute("BEGIN IMMEDIATE")
        conn.execute("DELETE FROM scores WHERE round_id = ?", (round_id,))
        conn.execute("DELETE FROM tables WHERE round_id = ?", (round_id,))
        plan, warnings = build_table_plan(conn, round_no)
        table_index = 1
        response_tables: list[dict] = []
        for item in plan:
            players = item["players"]
            table_no = f"R{round_no}-{table_index:02d}"
            player_ids = [player["id"] for player in players]
            cursor = conn.execute(
                """
                INSERT INTO tables(
                    round_id, table_no, age_group, game_type, player_ids_json,
                    is_bye, participant_count, has_free_agent
                ) VALUES(?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    round_id,
                    table_no,
                    item["age_group"],
                    item["game_type"],
                    json.dumps(player_ids, ensure_ascii=False),
                    int(item["is_bye"]),
                    len(players),
                    int(item["has_free_agent"]),
                ),
            )
            table_id = int(cursor.lastrowid)
            token = None
            if item["is_bye"]:
                token = make_table_token(table_id, round_id)
                conn.execute(
                    "UPDATE tables SET token = ?, token_created_at = ? WHERE id = ?",
                    (token, datetime.now().isoformat(timespec="seconds"), table_id),
                )
            else:
                token = make_table_token(table_id, round_id)
                conn.execute(
                    "UPDATE tables SET token = ?, token_created_at = ? WHERE id = ?",
                    (token, datetime.now().isoformat(timespec="seconds"), table_id),
                )
            response_tables.append(
                {
                    "id": table_id,
                    "table_no": table_no,
                    "age_group": item["age_group"],
                    "game_type": item["game_type"],
                    "players": player_public(players),
                    "is_bye": item["is_bye"],
                    "participant_count": len(players),
                    "has_free_agent": item["has_free_agent"],
                }
            )
            table_index += 1
        conn.execute("UPDATE rounds SET status = '进行中' WHERE id = ?", (round_id,))
        conn.execute("COMMIT")

    return {"round_no": round_no, "tables": response_tables, "warnings": warnings}


def qr_public_path(qr_file: Path) -> str:
    version = qr_file.stat().st_mtime_ns if qr_file.exists() else int(datetime.now().timestamp() * 1000)
    return f"/static/qrcodes/{qr_file.name}?v={version}"


def refresh_round_qrcodes(round_no: int, base_url: str) -> dict:
    from .security import make_table_token

    refreshed = 0
    with get_conn() as conn:
        round_id = get_round_id(conn, round_no)
        rows = conn.execute(
            """
            SELECT *
            FROM tables
            WHERE round_id = ? AND is_bye = 0
            ORDER BY id
            """,
            (round_id,),
        ).fetchall()
        conn.execute("BEGIN IMMEDIATE")
        for row in rows:
            if row["submitted_at"] is not None:
                continue
            token = make_table_token(row["id"], row["round_id"])
            conn.execute(
                "UPDATE tables SET token = ?, token_created_at = ? WHERE id = ?",
                (token, datetime.now().isoformat(timespec="seconds"), row["id"]),
            )
            refreshed += 1
        conn.execute("COMMIT")
    return {"round_no": round_no, "refreshed": refreshed, "tables": get_tables(round_no)}


def player_public(players: list[dict]) -> list[dict]:
    return [
        {
            "id": player["id"],
            "name": player["name"],
            "team_name": player["team_name"],
            "district": player["district"],
            "school": player["school"],
            "team_code": player["team_code"],
            "player_code": player["player_code"],
            "player_qr_token": player.get("player_qr_token"),
            "player_qr_path": player_qr_public_path(player["id"], player["name"], player["player_code"]),
            "age_group": player["age_group"],
            "game_type": player["game_type"],
            "seed_rank": player["seed_rank"],
            "identity_label": player["identity_label"],
        }
        for player in players
    ]


def player_status(token: str) -> dict:
    from .security import verify_player_token

    try:
        payload = verify_player_token(token)
    except ValueError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc
    with get_conn() as conn:
        if "player_code" in payload:
            player = conn.execute("SELECT * FROM players WHERE player_code = ?", (payload["player_code"],)).fetchone()
        else:
            player = conn.execute("SELECT * FROM players WHERE id = ?", (int(payload["player_id"]),)).fetchone()
        if not player:
            raise HTTPException(status_code=404, detail="选手不存在")
        player_dict = dict(player)
        if player_dict.get("player_qr_token") != token:
            raise HTTPException(status_code=403, detail="选手二维码已失效")
        player_id = int(player_dict["id"])

        score_rows = conn.execute(
            """
            SELECT r.round_no, r.status AS round_status, t.table_no, t.age_group AS table_age_group,
                   t.game_type AS table_game_type, t.is_bye, t.submitted_at,
                   s.table_rank, s.round_score, s.is_absent, s.is_advanced
            FROM scores s
            JOIN rounds r ON r.id = s.round_id
            JOIN tables t ON t.id = s.table_id
            WHERE s.player_id = ?
            ORDER BY r.round_no
            """,
            (player_id,),
        ).fetchall()
        table_rows = conn.execute(
            """
            SELECT t.*, r.round_no, r.status AS round_status
            FROM tables t
            JOIN rounds r ON r.id = t.round_id
            ORDER BY r.round_no, t.id
            """
        ).fetchall()
        team_rows = conn.execute(
            """
            SELECT COALESCE(SUM(s.round_score), 0) AS total_score,
                   COALESCE(SUM(CASE WHEN r.round_no IN (4,5,6) THEN s.round_score ELSE 0 END), 0) AS final_score
            FROM players p
            LEFT JOIN scores s ON s.player_id = p.id
            LEFT JOIN rounds r ON r.id = s.round_id
            WHERE p.team_name = ? AND p.game_type = ? AND p.identity_label != ?
            """,
            (player_dict["team_name"], player_dict["game_type"], IDENTITY_FREE_AGENT),
        ).fetchone()

    scores_by_round = {int(row["round_no"]): dict(row) for row in score_rows}
    score_history = []
    total_score = 0
    final_score = 0
    latest_score_round = 0
    latest_advanced: bool | None = None
    for round_no in range(1, 7):
        item = scores_by_round.get(round_no)
        if item:
            total_score += int(item["round_score"])
            if round_no in FINAL_ROUNDS:
                final_score += int(item["round_score"])
            latest_score_round = round_no
            latest_advanced = bool(item["is_advanced"])
        score_history.append(
            {
                "round_no": round_no,
                "table_no": item["table_no"] if item else "",
                "table_age_group": item["table_age_group"] if item else "",
                "game_type": item["table_game_type"] if item else player_dict["game_type"],
                "rank": item["table_rank"] if item else None,
                "score": item["round_score"] if item else None,
                "absent": bool(item["is_absent"]) if item else False,
                "advanced": bool(item["is_advanced"]) if item else None,
                "round_status": item["round_status"] if item else "未生成",
                "bye": bool(item["is_bye"]) if item else False,
            }
        )

    current_assignment = None
    for row in table_rows:
        ids = json.loads(row["player_ids_json"])
        if player_id not in ids:
            continue
        current_assignment = {
            "round_no": row["round_no"],
            "table_no": row["table_no"],
            "age_group": row["age_group"],
            "game_type": row["game_type"],
            "is_bye": bool(row["is_bye"]),
            "submitted": row["submitted_at"] is not None,
            "round_status": row["round_status"],
            "score_entry_url": f"/score?token={row['token']}" if row["token"] and row["submitted_at"] is None else "",
        }

    if current_assignment and current_assignment["submitted"] and latest_score_round == 6:
        next_message = "比赛已完成，请等待最终排名"
    elif current_assignment and current_assignment["submitted"] and latest_advanced:
        next_message = "本轮已晋级，等待管理员生成下一轮桌位"
    elif current_assignment and current_assignment["submitted"] and latest_advanced is False and latest_score_round in FINAL_ROUNDS:
        next_message = "决赛本轮未晋级"
    elif current_assignment and current_assignment["submitted"]:
        next_message = "本轮成绩已录入，等待下一轮分桌"
    elif current_assignment and current_assignment["is_bye"]:
        next_message = "本轮2人轮空晋级，请等待裁判录入轮空分"
    elif current_assignment:
        next_message = "请前往当前桌位比赛"
    elif latest_score_round and latest_advanced:
        next_message = "已晋级，等待管理员生成下一轮桌位"
    elif latest_score_round and latest_advanced is False and latest_score_round in PRELIM_ROUNDS:
        next_message = "预赛未晋级"
    elif latest_score_round and latest_advanced is False and latest_score_round in FINAL_ROUNDS:
        next_message = "决赛本轮未晋级"
    else:
        next_message = "等待分桌"

    return {
        "player": {
            **player_public([player_dict])[0],
            "total_score": total_score,
            "final_score": final_score,
        },
        "team_score": {
            "total_score": int(team_rows["total_score"] or 0),
            "final_score": int(team_rows["final_score"] or 0),
        },
        "score_history": score_history,
        "current_assignment": current_assignment,
        "message": next_message,
    }


def table_with_players(row: dict, players_by_id: dict[int, dict]) -> dict:
    ids = json.loads(row["player_ids_json"])
    players = [players_by_id[player_id] for player_id in ids if player_id in players_by_id]
    return {
        "id": row["id"],
        "round_id": row["round_id"],
        "table_no": row["table_no"],
        "age_group": row["age_group"],
        "game_type": row["game_type"],
        "players": player_public(players),
        "seat_no": row["seat_no"] if "seat_no" in row else None,
        "is_bye": bool(row["is_bye"]),
        "participant_count": row["participant_count"],
        "has_free_agent": bool(row["has_free_agent"]),
        "submitted": row["submitted_at"] is not None,
    }


def list_rounds() -> list[dict]:
    with get_conn() as conn:
        return rows_to_dicts(conn.execute("SELECT * FROM rounds ORDER BY round_no").fetchall())


def get_tables(round_no: int | None = None) -> list[dict]:
    with get_conn() as conn:
        if round_no is None:
            rows = conn.execute("SELECT t.* FROM tables t ORDER BY t.id").fetchall()
        else:
            rows = conn.execute(
                """
                SELECT t.*
                FROM tables t
                JOIN rounds r ON r.id = t.round_id
                WHERE r.round_no = ?
                ORDER BY t.id
                """,
                (round_no,),
            ).fetchall()
        table_rows = [dict(row) for row in rows]
        player_ids = sorted({player_id for row in table_rows for player_id in json.loads(row["player_ids_json"])})
        players_by_id: dict[int, dict] = {}
        if player_ids:
            placeholders = ",".join("?" for _ in player_ids)
            players = rows_to_dicts(conn.execute(f"SELECT * FROM players WHERE id IN ({placeholders})", tuple(player_ids)).fetchall())
            players_by_id = {int(player["id"]): player for player in players}
    return [table_with_players(row, players_by_id) for row in table_rows]


def venue_seat_positions() -> list[dict]:
    image_width = 1661
    image_height = 4456
    rows = [
        (["113", "112", "111", "110"], [810, 930, 1110, 1292], 272),
        (["109", "108", "107", "106"], [810, 930, 1110, 1292], 386),
        (["105", "104", "103", "102"], [810, 930, 1110, 1292], 505),
        (["101", "100", "099", "098"], [810, 930, 1110, 1292], 616),
        (["097", "096", "095", "094"], [810, 930, 1110, 1292], 733),
        (["093", "092", "091", "090"], [810, 930, 1110, 1292], 964),
        (["089", "088", "087", "086"], [810, 930, 1110, 1292], 1083),
        (["085", "084", "083", "082"], [810, 930, 1110, 1292], 1194),
        (["081", "080", "079", "078"], [810, 930, 1110, 1292], 1311),
        (["077", "076"], [383, 579], 964),
        (["075", "074", "073"], [279, 398, 579], 1082),
        (["072", "071", "070"], [279, 398, 579], 1194),
        (["069", "068", "067"], [279, 398, 579], 1311),
        (["066", "065"], [398, 579], 1431),
        (["064", "063"], [398, 579], 1549),
        (["062", "061", "060"], [279, 398, 579], 1660),
        (["059", "058", "057"], [279, 398, 579], 1777),
        (["056", "055", "054"], [279, 398, 579], 1901),
        (["053", "052"], [398, 579], 2020),
        (["051", "050"], [398, 579], 2132),
        (["049", "048"], [398, 579], 2249),
        (["047", "046"], [398, 579], 2368),
        (["045", "044"], [398, 579], 2486),
        (["043", "042"], [822, 941], 1681),
        (["041", "040"], [822, 941], 1798),
        (["039", "038"], [822, 941], 1921),
        (["037", "036"], [850, 1000], 2051),
        (["035", "034"], [996, 1115], 2205),
        (["033", "032", "031", "030"], [857, 976, 1163, 1348], 2395),
        (["029", "028", "027", "026"], [857, 976, 1163, 1348], 2518),
        (["025", "024", "023", "022"], [857, 976, 1163, 1348], 2640),
        (["021", "020", "019", "018"], [857, 976, 1163, 1348], 2768),
        (["017"], [1163], 2886),
        (["016", "015", "014", "013"], [857, 976, 1163, 1348], 3005),
        (["012", "011", "010", "009"], [857, 976, 1163, 1348], 3128),
        (["008", "007", "006", "005"], [857, 976, 1163, 1348], 3250),
        (["004", "003", "002", "001"], [857, 976, 1163, 1348], 3378),
    ]
    seats: list[dict] = []
    for seat_numbers, xs, y in rows:
        for seat_no, x in zip(seat_numbers, xs):
            seats.append({"seat_no": seat_no, "x": round(x / image_width * 100, 3), "y": round(y / image_height * 100, 3)})
    return sorted(seats, key=lambda item: item["seat_no"])


def attach_round_scores(tables: list[dict], round_no: int) -> list[dict]:
    player_ids = [player["id"] for table in tables for player in table["players"]]
    table_ids = [int(table["id"]) for table in tables]
    if not player_ids:
        return tables
    placeholders = ",".join("?" for _ in player_ids)
    table_placeholders = ",".join("?" for _ in table_ids)
    with get_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT s.player_id, s.table_rank, s.round_score, s.is_absent, s.is_advanced
            FROM scores s
            JOIN rounds r ON r.id = s.round_id
            WHERE r.round_no = ? AND s.player_id IN ({placeholders})
            """,
            (round_no, *player_ids),
        ).fetchall()
        submission_rows = conn.execute(
            f"""
            SELECT js.table_id, js.judge_name, js.submitted_at, js.scores_json,
                   j.account, j.claimed_by
            FROM judge_submissions js
            JOIN judges j ON j.id = js.judge_id
            WHERE js.table_id IN ({table_placeholders})
            ORDER BY js.submitted_at DESC, js.id DESC
            """,
            tuple(table_ids),
        ).fetchall()
    scores_by_player = {int(row["player_id"]): dict(row) for row in rows}
    submissions_by_table: dict[int, dict] = {}
    for row in submission_rows:
        table_id = int(row["table_id"])
        if table_id in submissions_by_table:
            continue
        item = dict(row)
        item["scores"] = json.loads(item.pop("scores_json") or "[]")
        submissions_by_table[table_id] = item
    for table in tables:
        table["judge_submission"] = submissions_by_table.get(int(table["id"]))
        for player in table["players"]:
            score = scores_by_player.get(int(player["id"]))
            player["round_result"] = {
                "rank": score["table_rank"] if score else None,
                "score": score["round_score"] if score else None,
                "is_absent": bool(score["is_absent"]) if score else False,
                "is_advanced": bool(score["is_advanced"]) if score else False,
                "submitted": score is not None,
            }
        if table["judge_submission"]:
            table["judge_submission"]["scores"] = [
                {
                    "player_id": int(player["id"]),
                    "score": scores_by_player[int(player["id"])]["round_score"],
                    "rank": scores_by_player[int(player["id"])]["table_rank"],
                    "absent": bool(scores_by_player[int(player["id"])]["is_absent"]),
                    "advanced": bool(scores_by_player[int(player["id"])]["is_advanced"]),
                }
                for player in table["players"]
                if int(player["id"]) in scores_by_player
            ]
    return tables


def table_sort_key_for_seating(table: dict, priority: str, group_counts: Counter) -> tuple:
    if priority == "table_no":
        return (table["table_no"],)
    if priority in AGE_GROUPS:
        return (0 if table["age_group"] == priority else 1, table["table_no"])
    if priority in GAMES:
        return (0 if table["game_type"] == priority else 1, table["table_no"])
    return (group_counts[table["age_group"]], table["age_group"], table["table_no"])


def arrange_tables_for_seating(tables: list[dict], seats: list[dict], priority: str) -> tuple[list[dict], list[dict]]:
    seat_by_no = {seat["seat_no"]: {**seat, "table": None, "manual": False} for seat in seats}
    manually_seated: set[int] = set()
    for table in tables:
        seat_no = table.get("seat_no")
        if seat_no in seat_by_no and seat_by_no[seat_no]["table"] is None:
            seat_by_no[seat_no]["table"] = table
            seat_by_no[seat_no]["manual"] = True
            manually_seated.add(int(table["id"]))

    group_counts = Counter(table["age_group"] for table in tables)
    remaining = [table for table in tables if int(table["id"]) not in manually_seated]
    remaining.sort(key=lambda table: table_sort_key_for_seating(table, priority, group_counts))
    for seat in seats:
        target = seat_by_no[seat["seat_no"]]
        if target["table"] is not None or not remaining:
            continue
        target["table"] = remaining.pop(0)

    assignments = [seat_by_no[seat["seat_no"]] for seat in seats]
    return assignments, remaining


def seating_chart(round_no: int, priority: str = "small_group_first") -> dict:
    tables = attach_round_scores(get_tables(round_no), round_no)
    seats = venue_seat_positions()
    assignments, overflow = arrange_tables_for_seating(tables, seats, priority)
    seat_map = BASE_DIR / "static" / "assets" / "seat_map.jpg"
    return {
        "round_no": round_no,
        "priority": priority,
        "map_image": "/static/assets/seat_map.jpg" if seat_map.exists() else "",
        "seat_capacity": len(seats),
        "table_count": len(tables),
        "occupied_count": min(len(tables), len(seats)),
        "overflow_count": len(overflow),
        "seats": assignments,
        "assignments": [item for item in assignments if item["table"]],
        "overflow": overflow,
    }


def advancement_chart() -> dict:
    round_titles = {
        1: "第一轮",
        2: "第二轮",
        3: "第三轮",
        4: "第四轮",
        5: "第五轮",
        6: "第六轮",
    }
    round_notes = {
        1: "每桌前2名晋级",
        2: "每桌前2名晋级",
        3: "每桌前2名晋级",
        4: "每桌第1名晋级",
        5: "按组别游戏取前4",
        6: "最终决赛",
    }
    groups: dict[tuple[str, str], dict] = {}
    chart_rounds = range(1, 7)
    for round_no in chart_rounds:
        tables = attach_round_scores(get_tables(round_no), round_no)
        for table in tables:
            key = (table["age_group"], table["game_type"])
            group = groups.setdefault(
                key,
                {
                    "key": f"{table['age_group']}__{table['game_type']}",
                    "age_group": table["age_group"],
                    "game_type": table["game_type"],
                    "rounds": {
                        item: {
                            "round_no": item,
                            "title": round_titles[item],
                            "note": round_notes[item],
                            "tables": [],
                        }
                        for item in chart_rounds
                    },
                },
            )
            players = sorted(
                table["players"],
                key=lambda player: (
                    player.get("round_result", {}).get("rank") or 999,
                    player.get("seed_rank") or 999999,
                ),
            )
            submission = table.get("judge_submission")
            group["rounds"][round_no]["tables"].append(
                {
                    "id": table["id"],
                    "table_no": table["table_no"],
                    "seat_no": table.get("seat_no"),
                    "submitted": table["submitted"],
                    "is_bye": table["is_bye"],
                    "submitted_at": submission.get("submitted_at") if submission else None,
                    "players": players,
                }
            )

    normalized_groups = []
    for group in groups.values():
        rounds = []
        for round_info in group["rounds"].values():
            round_info["tables"].sort(key=lambda table: table["table_no"])
            rounds.append(round_info)
        normalized_groups.append({**group, "rounds": rounds})
    normalized_groups.sort(key=lambda item: (AGE_GROUPS.index(item["age_group"]) if item["age_group"] in AGE_GROUPS else 99, item["game_type"]))
    return {"groups": normalized_groups}


def set_table_seat(round_no: int, table_id: int, seat_no: str | None) -> dict:
    valid_seats = {seat["seat_no"] for seat in venue_seat_positions()}
    normalized_seat = clean_cell(seat_no) if seat_no is not None else ""
    if normalized_seat and normalized_seat not in valid_seats:
        raise HTTPException(status_code=400, detail="座位号不存在")
    with get_conn() as conn:
        round_id = get_round_id(conn, round_no)
        table = conn.execute("SELECT id FROM tables WHERE id = ? AND round_id = ?", (table_id, round_id)).fetchone()
        if not table:
            raise HTTPException(status_code=404, detail="桌位不存在")
        conn.execute("BEGIN IMMEDIATE")
        if normalized_seat:
            conn.execute(
                "UPDATE tables SET seat_no = NULL WHERE round_id = ? AND seat_no = ? AND id != ?",
                (round_id, normalized_seat, table_id),
            )
            conn.execute("UPDATE tables SET seat_no = ? WHERE id = ?", (normalized_seat, table_id))
        else:
            conn.execute("UPDATE tables SET seat_no = NULL WHERE id = ?", (table_id,))
        conn.execute("COMMIT")
    return {"ok": True}


def judge_public(row: dict) -> dict:
    return {
        "id": row["id"],
        "account": row["account"],
        "initial_password": row["initial_password"],
        "claimed_by": row["claimed_by"] or "",
        "note": row["note"] or "",
        "is_active": bool(row["is_active"]),
        "created_at": row["created_at"],
        "last_login_at": row["last_login_at"],
        "submission_count": int(row.get("submission_count") or 0),
    }


def next_judge_account(conn) -> int:
    rows = conn.execute("SELECT account FROM judges WHERE account LIKE 'J%'").fetchall()
    maximum = 0
    for row in rows:
        suffix = str(row["account"])[1:]
        if suffix.isdigit():
            maximum = max(maximum, int(suffix))
    return maximum + 1


def generate_judge_accounts(count: int) -> dict:
    from .security import hash_password

    if count < 1 or count > 300:
        raise HTTPException(status_code=400, detail="一次生成数量请控制在 1-300 个")
    created = []
    now = datetime.now().isoformat(timespec="seconds")
    with get_conn() as conn:
        conn.execute("BEGIN IMMEDIATE")
        start = next_judge_account(conn)
        for offset in range(count):
            account = f"J{start + offset:03d}"
            password = "".join(secrets.choice("0123456789") for _ in range(6))
            conn.execute(
                """
                INSERT INTO judges(account, password_hash, initial_password, created_at)
                VALUES(?, ?, ?, ?)
                """,
                (account, hash_password(password), password, now),
            )
            created.append({"account": account, "password": password})
        conn.execute("COMMIT")
    return {"created": created, "text": judge_group_text(created)}


def judge_group_text(accounts: list[dict]) -> str:
    lines = ["裁判账号领取表", "请接龙：姓名 + 领取账号", ""]
    lines.extend(f"{item['account']} / {item['password']}" for item in accounts)
    return "\n".join(lines)


def list_judges() -> list[dict]:
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT j.*,
                   COUNT(js.id) AS submission_count
            FROM judges j
            LEFT JOIN judge_submissions js ON js.judge_id = j.id
            GROUP BY j.id
            ORDER BY j.account
            """
        ).fetchall()
    return [judge_public(dict(row)) for row in rows]


def update_judge(judge_id: int, payload: dict) -> dict:
    fields = []
    params: list[Any] = []
    if "claimed_by" in payload:
        fields.append("claimed_by = ?")
        params.append(clean_cell(payload.get("claimed_by")))
    if "note" in payload:
        fields.append("note = ?")
        params.append(clean_cell(payload.get("note")))
    if "is_active" in payload:
        fields.append("is_active = ?")
        params.append(1 if payload.get("is_active") else 0)
    if not fields:
        return {"ok": True}
    params.append(judge_id)
    with get_conn() as conn:
        conn.execute("BEGIN IMMEDIATE")
        conn.execute(f"UPDATE judges SET {', '.join(fields)} WHERE id = ?", params)
        conn.execute("COMMIT")
    return {"ok": True}


def reset_judge_password(judge_id: int) -> dict:
    from .security import hash_password

    password = "".join(secrets.choice("0123456789") for _ in range(6))
    with get_conn() as conn:
        row = conn.execute("SELECT account FROM judges WHERE id = ?", (judge_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="裁判账号不存在")
        conn.execute("BEGIN IMMEDIATE")
        conn.execute(
            "UPDATE judges SET password_hash = ?, initial_password = ? WHERE id = ?",
            (hash_password(password), password, judge_id),
        )
        conn.execute("COMMIT")
    return {"account": row["account"], "password": password}


def judge_login(account: str, password: str, judge_name: str) -> dict:
    from .security import make_judge_token, verify_password

    clean_name = clean_cell(judge_name)
    if not clean_name:
        raise HTTPException(status_code=400, detail="请填写裁判姓名")
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM judges WHERE account = ?", (clean_cell(account),)).fetchone()
        if not row or not verify_password(password, row["password_hash"]):
            raise HTTPException(status_code=401, detail="账号或密码错误")
        if not row["is_active"]:
            raise HTTPException(status_code=403, detail="该裁判账号已禁用")
        now = datetime.now().isoformat(timespec="seconds")
        conn.execute("UPDATE judges SET last_login_at = ? WHERE id = ?", (now, row["id"]))
        judge = dict(row)
        judge["last_login_at"] = now
    return {
        "token": make_judge_token(judge["id"], judge["account"], clean_name),
        "judge": {**judge_public({**judge, "submission_count": 0}), "login_name": clean_name},
    }


def judge_from_token(conn, token: str | None) -> dict:
    from .security import verify_judge_token

    if not token:
        raise HTTPException(status_code=401, detail="请先使用裁判账号登录")
    try:
        payload = verify_judge_token(token)
    except ValueError as exc:
        raise HTTPException(status_code=401, detail=str(exc)) from exc
    row = conn.execute("SELECT * FROM judges WHERE id = ?", (payload["judge_id"],)).fetchone()
    if not row:
        raise HTTPException(status_code=401, detail="裁判账号不存在")
    if not row["is_active"]:
        raise HTTPException(status_code=403, detail="该裁判账号已禁用")
    judge = dict(row)
    judge["login_name"] = clean_cell(payload.get("judge_name"))
    if not judge["login_name"]:
        raise HTTPException(status_code=401, detail="裁判登录已失效，请重新登录")
    return judge


def list_judge_submissions() -> list[dict]:
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT js.*, j.account, j.claimed_by, r.round_no, t.table_no, t.seat_no,
                   t.age_group, t.game_type
            FROM judge_submissions js
            JOIN judges j ON j.id = js.judge_id
            JOIN rounds r ON r.id = js.round_id
            JOIN tables t ON t.id = js.table_id
            ORDER BY js.submitted_at DESC, js.id DESC
            """
        ).fetchall()
        table_ids = sorted({int(row["table_id"]) for row in rows})
        current_scores_by_table: dict[int, list[dict]] = {}
        corrections_by_table: dict[int, list[dict]] = {}
        if table_ids:
            placeholders = ",".join("?" for _ in table_ids)
            score_rows = conn.execute(
                f"""
                SELECT s.table_id, s.player_id, s.table_rank, s.round_score, s.is_absent, s.is_advanced,
                       p.name, p.player_code, p.team_name
                FROM scores s
                JOIN players p ON p.id = s.player_id
                WHERE s.table_id IN ({placeholders})
                ORDER BY s.table_id, s.table_rank, s.player_id
                """,
                tuple(table_ids),
            ).fetchall()
            for score in score_rows:
                item = dict(score)
                # Keep the score submission API stable after loading corrected values from scores.
                item["rank"] = item.pop("table_rank")
                item["score"] = item.pop("round_score")
                item["absent"] = bool(item.pop("is_absent"))
                item["advanced"] = bool(item.pop("is_advanced"))
                current_scores_by_table.setdefault(int(item["table_id"]), []).append(item)

            correction_rows = conn.execute(
                f"""
                SELECT id, table_id, operator_name, reason, corrected_at, old_scores_json, new_scores_json
                FROM score_corrections
                WHERE table_id IN ({placeholders})
                ORDER BY corrected_at DESC, id DESC
                """,
                tuple(table_ids),
            ).fetchall()
            for correction in correction_rows:
                item = dict(correction)
                item["old_scores"] = json.loads(item.pop("old_scores_json") or "[]")
                item["new_scores"] = json.loads(item.pop("new_scores_json") or "[]")
                corrections_by_table.setdefault(int(item["table_id"]), []).append(item)

    items = []
    for row in rows_to_dicts(rows):
        table_id = int(row["table_id"])
        row["original_scores"] = json.loads(row.pop("scores_json") or "[]")
        row["scores"] = current_scores_by_table.get(table_id, row["original_scores"])
        row["correction_history"] = corrections_by_table.get(table_id, [])
        row["correction_count"] = len(row["correction_history"])
        items.append(row)
    return items


def score_page_data(token: str) -> dict:
    from .security import verify_token

    payload = verify_token(token)
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM tables WHERE id = ? AND round_id = ?", (payload["table_id"], payload["round_id"])).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="桌位不存在")
        if row["token"] != token:
            raise HTTPException(status_code=403, detail="二维码已失效")
        ids = json.loads(row["player_ids_json"])
        placeholders = ",".join("?" for _ in ids)
        players = rows_to_dicts(conn.execute(f"SELECT * FROM players WHERE id IN ({placeholders})", ids).fetchall())
        by_id = {player["id"]: player for player in players}
        round_no = conn.execute("SELECT round_no FROM rounds WHERE id = ?", (row["round_id"],)).fetchone()["round_no"]
    return {
        "valid": row["submitted_at"] is None,
        "submitted": row["submitted_at"] is not None,
        "round_no": round_no,
        "table_no": row["table_no"],
        "age_group": row["age_group"],
        "game_type": row["game_type"],
        "has_free_agent": bool(row["has_free_agent"]),
        "is_bye": bool(row["is_bye"]),
        "players": player_public([by_id[player_id] for player_id in ids if player_id in by_id]),
    }


def calculate_table_scores(table: Any, results: list[dict]) -> tuple[list[int], dict[int, dict], dict[int, int]]:
    expected_ids = [int(player_id) for player_id in json.loads(table["player_ids_json"])]
    result_by_id = {int(item["player_id"]): item for item in results}
    if set(result_by_id.keys()) != set(expected_ids):
        raise HTTPException(status_code=400, detail="提交选手名单与桌位不一致")

    present = [item for item in result_by_id.values() if not item.get("absent")]
    allowed_scores = set(SCORE_MAP[4])
    submitted_scores = [int(item["score"]) for item in present if item.get("score") is not None]
    if len(submitted_scores) != len(present):
        raise HTTPException(status_code=400, detail="请为所有未缺席选手选择分数")
    if any(score not in allowed_scores for score in submitted_scores):
        raise HTTPException(status_code=400, detail="分数只能选择：5、3、2、1；缺席为0分")

    ranked_present = sorted(
        present,
        key=lambda item: (-int(item["score"]), expected_ids.index(int(item["player_id"]))),
    )
    rank_by_id = {int(item["player_id"]): index + 1 for index, item in enumerate(ranked_present)}
    absent_ids = [player_id for player_id in expected_ids if result_by_id[player_id].get("absent")]
    for index, player_id in enumerate(absent_ids):
        rank_by_id[player_id] = len(present) + index + 1
    return expected_ids, result_by_id, rank_by_id


def table_score_snapshot(conn: Any, table: Any) -> list[dict]:
    expected_ids = [int(player_id) for player_id in json.loads(table["player_ids_json"])]
    rows = conn.execute(
        """
        SELECT s.player_id, s.table_rank, s.round_score, s.is_absent, s.is_advanced,
               p.name, p.player_code, p.team_name
        FROM scores s
        JOIN players p ON p.id = s.player_id
        WHERE s.table_id = ?
        """,
        (table["id"],),
    ).fetchall()
    by_id = {int(row["player_id"]): dict(row) for row in rows}
    snapshot = []
    for player_id in expected_ids:
        row = by_id.get(player_id)
        if not row:
            continue
        snapshot.append(
            {
                "player_id": player_id,
                "name": row["name"],
                "player_code": row["player_code"],
                "team_name": row["team_name"],
                "rank": row["table_rank"],
                "score": row["round_score"],
                "absent": bool(row["is_absent"]),
                "advanced": bool(row["is_advanced"]),
            }
        )
    return snapshot


def submit_score(payload: dict, client_ip: str = "", user_agent: str = "") -> dict:
    from .security import verify_token

    token_payload = verify_token(payload["token"])
    with get_conn() as conn:
        conn.execute("BEGIN IMMEDIATE")
        judge = judge_from_token(conn, payload.get("judge_token"))
        table = conn.execute(
            "SELECT * FROM tables WHERE id = ? AND round_id = ?",
            (token_payload["table_id"], token_payload["round_id"]),
        ).fetchone()
        if not table:
            raise HTTPException(status_code=404, detail="桌位不存在")
        if table["token"] != payload["token"] or table["submitted_at"] is not None:
            raise HTTPException(status_code=409, detail="本桌成绩已录入，不可重复提交")

        expected_ids, result_by_id, rank_by_id = calculate_table_scores(table, payload["results"])

        round_no = conn.execute("SELECT round_no FROM rounds WHERE id = ?", (table["round_id"],)).fetchone()["round_no"]
        for player_id in expected_ids:
            item = result_by_id[player_id]
            absent = bool(item.get("absent"))
            rank = rank_by_id[player_id]
            score = 0 if absent else int(item["score"])
            advanced = is_player_advanced(round_no, int(table["participant_count"]), bool(table["is_bye"]), rank, absent)
            insert_score(conn, player_id, table["round_id"], table["id"], rank, score, absent, advanced)

        submitted_at = datetime.now().isoformat(timespec="seconds")
        conn.execute(
            "UPDATE tables SET submitted_at = ?, token = NULL WHERE id = ?",
            (submitted_at, table["id"]),
        )
        conn.execute(
            """
            INSERT INTO judge_submissions(
                judge_id, round_id, table_id, judge_name, submitted_at, client_ip, user_agent, scores_json
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                judge["id"],
                table["round_id"],
                table["id"],
                judge["login_name"],
                submitted_at,
                client_ip,
                user_agent,
                json.dumps(
                    [
                        {
                            "player_id": player_id,
                            "rank": rank_by_id[player_id],
                            "score": 0 if result_by_id[player_id].get("absent") else int(result_by_id[player_id]["score"]),
                            "absent": bool(result_by_id[player_id].get("absent")),
                            "advanced": is_player_advanced(
                                round_no,
                                int(table["participant_count"]),
                                bool(table["is_bye"]),
                                rank_by_id[player_id],
                                bool(result_by_id[player_id].get("absent")),
                            ),
                        }
                        for player_id in expected_ids
                    ],
                    ensure_ascii=False,
                ),
            ),
        )
        if not conn.execute("SELECT 1 FROM tables WHERE round_id = ? AND submitted_at IS NULL", (table["round_id"],)).fetchone():
            if round_no == 5:
                recompute_round5_advancement(conn, table["round_id"])
            conn.execute("UPDATE rounds SET status = '已结束' WHERE id = ?", (table["round_id"],))
        conn.execute("COMMIT")
    return {"ok": True, "message": "成绩提交成功"}


def correct_score(round_no: int, table_id: int, payload: dict) -> dict:
    operator_name = clean_cell(payload.get("operator_name"))
    reason = clean_cell(payload.get("reason"))
    if not operator_name or not reason:
        raise HTTPException(status_code=400, detail="请填写更正人和更正原因")

    with get_conn() as conn:
        conn.execute("BEGIN IMMEDIATE")
        round_id = get_round_id(conn, round_no)
        table = conn.execute("SELECT * FROM tables WHERE id = ? AND round_id = ?", (table_id, round_id)).fetchone()
        if not table:
            raise HTTPException(status_code=404, detail="桌位不存在")
        if table["submitted_at"] is None:
            raise HTTPException(status_code=400, detail="本桌尚未提交成绩，不能进行更正")
        if round_no < 6:
            next_round_id = get_round_id(conn, round_no + 1)
            next_round_has_tables = conn.execute("SELECT 1 FROM tables WHERE round_id = ? LIMIT 1", (next_round_id,)).fetchone()
            if next_round_has_tables:
                raise HTTPException(
                    status_code=409,
                    detail=f"第{round_no + 1}轮桌位已生成。为保护后续晋级，本次更正已被阻止；请在生成下一轮前完成成绩更正。",
                )

        expected_ids, result_by_id, rank_by_id = calculate_table_scores(table, payload["results"])
        old_scores = table_score_snapshot(conn, table)
        if len(old_scores) != len(expected_ids):
            raise HTTPException(status_code=400, detail="本桌成绩数据不完整，无法更正")

        for player_id in expected_ids:
            item = result_by_id[player_id]
            absent = bool(item.get("absent"))
            rank = rank_by_id[player_id]
            score = 0 if absent else int(item["score"])
            advanced = is_player_advanced(round_no, int(table["participant_count"]), bool(table["is_bye"]), rank, absent)
            insert_score(conn, player_id, round_id, table_id, rank, score, absent, advanced)

        if round_no == 5:
            recompute_round5_advancement(conn, round_id)
        new_scores = table_score_snapshot(conn, table)
        original_submission = conn.execute(
            "SELECT id FROM judge_submissions WHERE table_id = ? ORDER BY id ASC LIMIT 1",
            (table_id,),
        ).fetchone()
        corrected_at = datetime.now().isoformat(timespec="seconds")
        conn.execute(
            """
            INSERT INTO score_corrections(
                round_id, table_id, original_judge_submission_id, operator_name, reason,
                corrected_at, old_scores_json, new_scores_json
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                round_id,
                table_id,
                original_submission["id"] if original_submission else None,
                operator_name,
                reason,
                corrected_at,
                json.dumps(old_scores, ensure_ascii=False),
                json.dumps(new_scores, ensure_ascii=False),
            ),
        )
        conn.execute("COMMIT")
    return {"ok": True, "message": "成绩已更正，积分榜和晋级状态已同步更新", "corrected_at": corrected_at}


def leaderboard(limit: int = 0) -> list[dict]:
    limit_sql = "LIMIT ?" if limit and limit > 0 else ""
    params: tuple[int, ...] = (limit,) if limit and limit > 0 else ()
    with get_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT p.id, p.name, p.team_name, p.district, p.school, p.team_code, p.player_code,
                   p.age_group, p.game_type, p.seed_rank, p.identity_label,
                   COALESCE(SUM(s.round_score), 0) AS total_score,
                   COALESCE(SUM(CASE WHEN r.round_no IN (4,5,6) THEN s.round_score ELSE 0 END), 0) AS final_score
            FROM players p
            LEFT JOIN scores s ON s.player_id = p.id
            LEFT JOIN rounds r ON r.id = s.round_id
            WHERE p.is_active = 1
            GROUP BY p.id
            ORDER BY total_score DESC, final_score DESC, p.seed_rank ASC
            {limit_sql}
            """,
            params,
        ).fetchall()
    return rows_to_dicts(rows)


def team_rankings(age_group: str | None = None, game_type: str | None = None, metric: str = "final") -> list[dict]:
    filters = ["p.is_active = 1", "p.identity_label != ?"]
    params: list[Any] = [IDENTITY_FREE_AGENT]
    if age_group:
        filters.append("p.age_group = ?")
        params.append(age_group)
    if game_type:
        filters.append("p.game_type = ?")
        params.append(game_type)
    where_sql = " AND ".join(filters)

    with get_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT p.id, p.name, p.team_name, p.district, p.school, p.team_code, p.player_code,
                   p.age_group, p.game_type, p.seed_rank,
                   COALESCE(SUM(s.round_score), 0) AS total_score,
                   COALESCE(SUM(CASE WHEN r.round_no IN (4,5,6) THEN s.round_score ELSE 0 END), 0) AS final_score
            FROM players p
            LEFT JOIN scores s ON s.player_id = p.id
            LEFT JOIN rounds r ON r.id = s.round_id
            WHERE {where_sql}
            GROUP BY p.id
            ORDER BY p.seed_rank ASC
            """,
            params,
        ).fetchall()

    grouped: dict[tuple[str, str, str], dict] = {}
    for row in rows_to_dicts(rows):
        key = (row["team_code"] or row["team_name"], row["school"], row["age_group"], row["game_type"])
        team = grouped.setdefault(
            key,
            {
                "team_name": row["team_name"],
                "district": row["district"],
                "school": row["school"],
                "team_code": row["team_code"],
                "age_group": row["age_group"],
                "game_type": row["game_type"],
                "member_count": 0,
                "total_score": 0,
                "final_score": 0,
                "seed_sum": 0,
                "members": [],
            },
        )
        team["member_count"] += 1
        team["total_score"] += int(row["total_score"])
        team["final_score"] += int(row["final_score"])
        team["seed_sum"] += int(row["seed_rank"])
        team["members"].append({"name": row["name"], "player_code": row["player_code"]})

    score_key = "final_score" if metric == "final" else "total_score"
    other_key = "total_score" if score_key == "final_score" else "final_score"
    teams = list(grouped.values())
    for team in teams:
        team["avg_seed"] = team["seed_sum"] / max(team["member_count"], 1)
    teams.sort(key=lambda item: (-item[score_key], -item[other_key], item["avg_seed"], item["team_name"]))
    for index, team in enumerate(teams, start=1):
        team["seed"] = index
        team["score"] = team[score_key]
    return teams


def next_power_of_two(value: int) -> int:
    size = 1
    while size < max(value, 2):
        size *= 2
    return size


def seed_slot_order(size: int) -> list[int]:
    order = [1, 2]
    while len(order) < size:
        mirror = len(order) * 2 + 1
        order = [item for seed in order for item in (seed, mirror - seed)]
    return order


def round_name(size: int, round_index: int, total_rounds: int) -> str:
    if round_index == total_rounds - 1:
        return "总决赛"
    if round_index == total_rounds - 2:
        return "半决赛"
    if round_index == total_rounds - 3:
        return "四分之一决赛"
    teams_left = size // (2**round_index)
    return f"{teams_left}强"


def build_team_bracket(age_group: str | None = None, game_type: str | None = None, metric: str = "final") -> dict:
    teams = team_rankings(age_group, game_type, metric)
    bracket_size = next_power_of_two(len(teams))
    team_by_seed = {team["seed"]: team for team in teams}
    slots = [team_by_seed.get(seed) for seed in seed_slot_order(bracket_size)]
    total_rounds = bracket_size.bit_length() - 1
    rounds: list[dict] = []
    current_slots: list[dict | None] = slots

    for round_index in range(total_rounds):
        matches = []
        next_slots: list[dict | None] = []
        for match_index in range(0, len(current_slots), 2):
            left = current_slots[match_index]
            right = current_slots[match_index + 1]
            if left and not right:
                next_slots.append(left)
            elif right and not left:
                next_slots.append(right)
            else:
                next_slots.append(None)
            matches.append(
                {
                    "match_no": match_index // 2 + 1,
                    "left": left,
                    "right": right,
                    "bye": bool((left and not right) or (right and not left)),
                }
            )
        rounds.append({"name": round_name(bracket_size, round_index, total_rounds), "matches": matches})
        current_slots = next_slots

    return {
        "team_count": len(teams),
        "bracket_size": bracket_size,
        "metric": metric,
        "age_group": age_group,
        "game_type": game_type,
        "rankings": teams,
        "rounds": rounds,
    }


def pool_status() -> dict:
    players = list_players()
    free_by_game: dict[str, int] = {game: 0 for game in GAMES}
    transition_count = 0
    for player in players:
        if player["is_transition"]:
            transition_count += 1
        if player["identity_label"] == IDENTITY_FREE_AGENT:
            free_by_game[player["game_type"]] += 1
    return {"transition_count": transition_count, "free_by_game": free_by_game}


def copy_cell_format(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)


def prepare_live_score_sheet(ws, required_rows: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= 3:
            ws.unmerge_cells(str(merged_range))

    max_col = max(ws.max_column, 12)
    required_max_row = max(3 + required_rows, ws.max_row)
    for row in range(3, required_max_row + 1):
        ws.row_dimensions[row].height = ws.row_dimensions[3].height or 28
        for col in range(1, max_col + 1):
            copy_cell_format(ws.cell(3, col), ws.cell(row, col))
            ws.cell(row, col).value = None

    if ws.max_column < 12:
        for row in range(1, required_max_row + 1):
            copy_cell_format(ws.cell(row, 11), ws.cell(row, 12))

    try:
        ws.unmerge_cells("A1:L1")
    except ValueError:
        pass
    ws.merge_cells("A1:L1")
    ws["L2"] = "所属游戏"
    copy_cell_format(ws["K2"], ws["L2"])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    for column in "EFGHIJ":
        ws.column_dimensions[column].width = 10
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 12


def create_live_score_workbook() -> Workbook:
    wb = Workbook()
    wb.active.title = AGE_GROUPS[0]
    for age_group in AGE_GROUPS[1:]:
        wb.create_sheet(age_group)

    headers = [
        "区",
        "学校",
        "队伍名称",
        "队员",
        "第1轮",
        "第2轮",
        "第3轮",
        "第4轮",
        "第5轮",
        "第6轮",
        "团队总分",
        "所属游戏",
    ]
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for ws in wb.worksheets:
        ws.merge_cells("A1:L1")
        ws["A1"] = f"{ws.title}成绩表（实时）"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 24
        ws.row_dimensions[3].height = 32
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(2, col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            data_cell = ws.cell(3, col)
            data_cell.border = thin_border
            data_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return wb


def merge_if_needed(ws, start_row: int, end_row: int, columns: list[int]) -> None:
    if end_row <= start_row:
        return
    for col in columns:
        ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)


def player_score_rows() -> tuple[list[dict], dict[int, dict[int, int]]]:
    with get_conn() as conn:
        players = rows_to_dicts(
            conn.execute(
                """
                SELECT id, name, team_name, district, school, team_code, player_code,
                       age_group, game_type, seed_rank, identity_label
                FROM players
                WHERE is_active = 1
                ORDER BY age_group, seed_rank
                """
            ).fetchall()
        )
        scores = conn.execute(
            """
            SELECT s.player_id, r.round_no, s.round_score
            FROM scores s
            JOIN rounds r ON r.id = s.round_id
            """
        ).fetchall()

    score_map: dict[int, dict[int, int]] = defaultdict(dict)
    for row in scores:
        score_map[int(row["player_id"])][int(row["round_no"])] = int(row["round_score"])
    return players, score_map


def live_team_key(player: dict) -> tuple:
    if player.get("identity_label") == IDENTITY_FREE_AGENT:
        return ("free", player["id"])
    return (
        player.get("team_code") or "",
        player.get("district") or "",
        player.get("school") or "",
        player.get("team_name") or "",
        player.get("game_type") or "",
    )


def export_live_scoreboard() -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    output = EXPORT_DIR / f"实时成绩统计表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    players, score_map = player_score_rows()
    players_by_group: dict[str, list[dict]] = defaultdict(list)
    for player in players:
        players_by_group[player["age_group"]].append(player)

    wb = load_workbook(LIVE_SCORE_TEMPLATE) if LIVE_SCORE_TEMPLATE.exists() else create_live_score_workbook()
    for age_group in AGE_GROUPS:
        if age_group not in wb.sheetnames:
            wb.create_sheet(age_group)
        ws = wb[age_group]
        group_players = players_by_group.get(age_group, [])
        teams: dict[tuple, list[dict]] = defaultdict(list)
        for player in group_players:
            teams[live_team_key(player)].append(player)

        sorted_teams = sorted(teams.values(), key=lambda team: min(int(player["seed_rank"]) for player in team))
        required_rows = max(sum(len(team) for team in sorted_teams), 1)
        prepare_live_score_sheet(ws, required_rows)
        ws["A1"] = f"{age_group}成绩表（实时）"

        row_no = 3
        for team in sorted_teams:
            team.sort(key=lambda player: int(player["seed_rank"]))
            start_row = row_no
            team_total = 0
            for player in team:
                total = sum(score_map.get(int(player["id"]), {}).get(round_no, 0) for round_no in range(1, 7))
                team_total += total
                ws.cell(row_no, 1).value = player.get("district") or ""
                ws.cell(row_no, 2).value = player.get("school") or ""
                ws.cell(row_no, 3).value = player.get("team_name") or ""
                player_code = player.get("player_code") or f"ID{player['id']}"
                ws.cell(row_no, 4).value = f"{player_code}\n{player['name']}"
                ws.cell(row_no, 4).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                for round_no in range(1, 7):
                    value = score_map.get(int(player["id"]), {}).get(round_no)
                    ws.cell(row_no, 4 + round_no).value = value if value is not None else ""
                ws.cell(row_no, 12).value = player.get("game_type") or ""
                row_no += 1

            end_row = row_no - 1
            ws.cell(start_row, 11).value = team_total
            merge_if_needed(ws, start_row, end_row, [1, 2, 3, 11, 12])

        for row in ws.iter_rows(min_row=1, max_row=max(row_no - 1, 3), max_col=12):
            for cell in row:
                cell.alignment = copy(cell.alignment)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=cell.alignment.wrap_text,
                )
        ws.freeze_panes = "A3"

    wb.save(output)
    return output


def export_report() -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    output = EXPORT_DIR / f"桌游比赛最终报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with get_conn() as conn:
        detail = pd.read_sql_query(
            """
            SELECT p.player_code AS 队员编号, p.name AS 姓名, p.team_name AS 队伍名称,
                   p.team_code AS 队伍编号, p.district AS 区, p.school AS 学校,
                   p.age_group AS 年级组别, p.game_type AS 所属游戏,
                   p.seed_rank AS 赛前种子排名, p.identity_label AS 选手身份,
                   r.round_no AS 轮次, t.table_no AS 桌号, t.submitted_at AS 成绩提交时间, s.table_rank AS 桌内排名,
                   s.round_score AS 本轮得分, s.is_absent AS 是否缺席, s.is_advanced AS 是否晋级
            FROM players p
            LEFT JOIN scores s ON s.player_id = p.id
            LEFT JOIN rounds r ON r.id = s.round_id
            LEFT JOIN tables t ON t.id = s.table_id
            ORDER BY p.seed_rank, r.round_no
            """,
            conn,
        )
        personal = pd.read_sql_query(
            """
            SELECT p.player_code AS 队员编号, p.name AS 姓名, p.team_name AS 队伍名称,
                   p.team_code AS 队伍编号, p.district AS 区, p.school AS 学校,
                   p.age_group AS 年级组别, p.game_type AS 所属游戏,
                   p.seed_rank AS 赛前种子排名, p.identity_label AS 选手身份,
                   COALESCE(SUM(s.round_score), 0) AS 预赛加决赛总积分,
                   COALESCE(SUM(CASE WHEN r.round_no IN (4,5,6) THEN s.round_score ELSE 0 END), 0) AS 决赛单独积分
            FROM players p
            LEFT JOIN scores s ON s.player_id = p.id
            LEFT JOIN rounds r ON r.id = s.round_id
            WHERE p.is_active = 1
            GROUP BY p.id
            ORDER BY 决赛单独积分 DESC, 预赛加决赛总积分 DESC, p.seed_rank ASC
            """,
            conn,
        )
        team = pd.read_sql_query(
            """
            SELECT p.team_name AS 队伍名称, p.team_code AS 队伍编号,
                   p.district AS 区, p.school AS 学校, p.game_type AS 所属游戏,
                   SUM(CASE WHEN r.round_no IN (4,5,6) THEN s.round_score ELSE 0 END) AS 决赛团队总积分,
                   COUNT(DISTINCT p.id) AS 计入团队人数
            FROM players p
            LEFT JOIN scores s ON s.player_id = p.id
            LEFT JOIN rounds r ON r.id = s.round_id
            WHERE p.is_active = 1 AND p.identity_label != '自由人'
            GROUP BY p.team_name, p.team_code, p.district, p.school, p.game_type
            ORDER BY 决赛团队总积分 DESC, 队伍名称 ASC
            """,
            conn,
        )

    for frame in (detail, personal, team):
        for column in ["是否缺席", "是否晋级"]:
            if column in frame.columns:
                frame[column] = frame[column].map({1: "是", 0: "否"}).fillna("")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        detail.to_excel(writer, sheet_name="全量选手每轮明细", index=False)
        personal.to_excel(writer, sheet_name="最终个人排名", index=False)
        team.to_excel(writer, sheet_name="团队奖项排名", index=False)
        for worksheet in writer.book.worksheets:
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_cells in worksheet.columns:
                width = max(len(str(cell.value or "")) for cell in column_cells) + 4
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 32)
    return output
